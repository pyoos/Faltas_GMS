import os
import pandas as pd
import re
import random
import logging

from matplotlib.backends.backend_qt5agg import FigureCanvasQTAgg as FigureCanvas
from matplotlib.figure import Figure
import matplotlib.pyplot as plt

from openpyxl import Workbook
from openpyxl.styles import PatternFill

from PyQt5.QtGui import QColor
from PyQt5.QtWidgets import (
    QDialog, QFileDialog, QVBoxLayout, QTabWidget, QTableWidget, QTableWidgetItem, 
    QMessageBox, QLabel, QHBoxLayout, QHeaderView, QDateEdit, QPushButton, QLineEdit, QComboBox, QInputDialog, QListWidget, QApplication, QScrollArea, QWidget
)
from PyQt5.QtCore import Qt, QDate

#list of libraries needed to install
#pip install openpyxl
#pip install xlsxwriter
#there could be others

class ExcelHandler:
    def __init__(self, parent, grant_management, save_directory="uploaded_files"):
        self.parent = parent
        self.grant_management = grant_management
        self.group_color_mapping = {}  # Store group-value-to-color mapping
        self.total_cost = 0
        self.selected_sum_label = None
        self.sheet_data = None
        self.saved_excel_sheets = {}  # Dictionary to store saved Excel sheets
        self.save_directory = save_directory
        os.makedirs(self.save_directory, exist_ok=True)

    def upload_excel(self):
        options = QFileDialog.Options()
        options |= QFileDialog.ReadOnly
        file_path, _ = QFileDialog.getOpenFileName(
            self.parent, "Upload Inventory Excel File", "", "Excel Files (*.xlsx);;All Files (*)", options=options
        )

        if file_path:
            try:
                excel_data = pd.read_excel(file_path, sheet_name=None)
                excel_data = {name: data for name, data in excel_data.items() if not data.empty}

                if excel_data:
                    self.sheet_data = excel_data[list(excel_data.keys())[0]]  # Default to the first sheet
                    print("Excel Data Loaded:", self.sheet_data.head())  # Debugging statement
                    self.display_excel_contents(excel_data)
                else:
                    self.sheet_data = None  # Clear any previous data
                    QMessageBox.warning(self.parent, "No Data", "The uploaded Excel file contains no data.")
            except Exception as e:
                self.sheet_data = None  # Clear any previous data
                QMessageBox.critical(self.parent, "Error", f"An error occurred: {str(e)}")


    



    def save_grouped_data_with_highlights(self, grouped_data, title, column_name):
        """
        Save grouped data with row colors into an Excel file.
        """
        try:
            # Prompt user for save location
            options = QFileDialog.Options()
            file_path, _ = QFileDialog.getSaveFileName(
                self.parent,
                "Save Grouped Data As",
                f"grouped_data_{title.replace(' ', '_').lower()}.xlsx",
                "Excel Files (*.xlsx);;All Files (*)",
                options=options
            )
            if not file_path:
                return

            # Create Excel workbook and worksheet
            workbook = Workbook()
            sheet = workbook.active
            sheet.title = "Grouped Data"

            # Write column headers
            for col_num, col_name in enumerate(grouped_data.columns, start=1):
                sheet.cell(row=1, column=col_num, value=col_name)

            # Write data with colors
            for row_num, row_data in enumerate(grouped_data.itertuples(index=False), start=2):
                group_value = getattr(row_data, column_name)
                color_hex = self.group_color_mapping.get(group_value, "#FFFFFF").lstrip("#")  # Default white
                fill = PatternFill(start_color=color_hex, end_color=color_hex, fill_type="solid")

                for col_num, value in enumerate(row_data, start=1):
                    cell = sheet.cell(row=row_num, column=col_num, value=value)
                    cell.fill = fill

            # Save the workbook
            workbook.save(file_path)
            QMessageBox.information(self.parent, "Success", f"Grouped data saved to:\n{file_path}")

        except Exception as e:
            QMessageBox.critical(self.parent, "Error", f"An error occurred: {str(e)}")


    def add_data_to_sheet(self):
        """Add a blank row to the current sheet data."""
        if self.sheet_data is None:
            QMessageBox.warning(self.parent, "No Data", "No data has been loaded. Please upload an Excel file first.")
            return

        try:
            # Create a new row as a DataFrame with default empty values
            new_row = pd.DataFrame([{col: "" for col in self.sheet_data.columns}])
            
            # Use pd.concat to append the new row
            self.sheet_data = pd.concat([self.sheet_data, new_row], ignore_index=True)

            QMessageBox.information(self.parent, "Data Added", "A new row has been added to the dataset.")
        except Exception as e:
            QMessageBox.critical(self.parent, "Error", f"An error occurred while adding data: {str(e)}")

    def remove_data_from_sheet(self):
        """Remove a specific row or all rows from the current sheet data."""
        if self.sheet_data is None:
            QMessageBox.warning(self.parent, "No Data", "No data has been loaded. Please upload an Excel file first.")
            return

        # Create the dialog for row removal
        dialog = QDialog(self.parent)
        dialog.setWindowTitle("Remove Data")
        dialog.setStyleSheet("background-color: #cce7ff;")
        dialog.resize(400, 200)

        layout = QVBoxLayout()

        # Instructions Label
        instructions_label = QLabel("Enter the row index to remove, or click 'Delete All' to clear all data.")
        instructions_label.setStyleSheet("font-size: 16px; color: black;")
        layout.addWidget(instructions_label)

        # Row Index Input
        row_index_input = QInputDialog()
        row_index_input.setStyleSheet("font-size: 14px; color: black;")
        row_index_input_label = QLabel("Row Index:")
        row_index_input_label.setStyleSheet("font-size: 16px; color: black;")
        layout.addWidget(row_index_input_label)

        row_index_field = QLineEdit()
        row_index_field.setStyleSheet("font-size: 14px; color: black;")
        layout.addWidget(row_index_field)

        # Remove Row Button
        remove_row_button = QPushButton("Remove Row")
        remove_row_button.setStyleSheet("font-size: 16px; color: white; background-color: #F44336;")
        layout.addWidget(remove_row_button)

        # Delete All Button
        delete_all_button = QPushButton("Delete All")
        delete_all_button.setStyleSheet("font-size: 16px; color: white; background-color: #D32F2F;")
        layout.addWidget(delete_all_button)

        # Button Connections
        remove_row_button.clicked.connect(lambda: self.remove_row(dialog, row_index_field.text()))
        delete_all_button.clicked.connect(lambda: self.delete_all_rows(dialog))

        dialog.setLayout(layout)
        dialog.exec_()

    def remove_row(self, dialog, row_index):
        """Remove a specific row by index."""
        try:
            row_index = int(row_index)
            if 0 <= row_index < len(self.sheet_data):
                self.sheet_data = self.sheet_data.drop(index=row_index).reset_index(drop=True)
                QMessageBox.information(self.parent, "Row Removed", f"Row {row_index} has been removed.")
                dialog.accept()
            else:
                QMessageBox.warning(self.parent, "Invalid Index", "The entered row index is out of range.")
        except ValueError:
            QMessageBox.warning(self.parent, "Invalid Input", "Please enter a valid row index.")
        except Exception as e:
            QMessageBox.critical(self.parent, "Error", f"An error occurred while removing the row: {str(e)}")

    def delete_all_rows(self, dialog):
        """Clear all rows from the current sheet data."""
        confirmation = QMessageBox.question(
            self.parent, "Confirm Delete All", "Are you sure you want to delete all data?",
            QMessageBox.Yes | QMessageBox.No, QMessageBox.No
        )
        if confirmation == QMessageBox.Yes:
            self.sheet_data = self.sheet_data.iloc[0:0]  # Clear all rows
            QMessageBox.information(self.parent, "All Data Deleted", "All rows have been removed from the dataset.")
            dialog.accept()


    def visualize_data(self):
        """Visualize the dataset using matplotlib and embed the graph into the GUI."""
        if self.sheet_data is None or self.sheet_data.empty:
            QMessageBox.warning(self.parent, "No Data", "No data available to visualize. Please add or load data first.")
            return

        try:
            # Identify numeric columns
            numeric_columns = self.sheet_data.select_dtypes(include=["number"]).columns.tolist()
            
            if not numeric_columns:
                QMessageBox.warning(self.parent, "No Numeric Columns", "No numeric columns are available for visualization.")
                return

            # Ask user to select x and y columns
            x_column, ok_x = QInputDialog.getItem(self.parent, "Select X-axis", 
                                                "Choose a numeric column for the X-axis:", numeric_columns, 0, False)
            if not ok_x:
                return

            y_column, ok_y = QInputDialog.getItem(self.parent, "Select Y-axis", 
                                                "Choose a numeric column for the Y-axis:", numeric_columns, 0, False)
            if not ok_y:
                return

            # Clear any previous graph canvas
            if hasattr(self, 'graph_canvas'):
                self.graph_canvas.deleteLater()

            # Create a new figure and canvas
            self.graph_canvas = FigureCanvas(Figure(figsize=(6, 4)))
            ax = self.graph_canvas.figure.add_subplot(111)

            # Plot the data
            ax.plot(self.sheet_data[x_column], self.sheet_data[y_column], 
                    marker='o', linestyle='-', color='blue')
            ax.set_title(f"{y_column} vs {x_column}")
            ax.set_xlabel(x_column)
            ax.set_ylabel(y_column)
            ax.grid(True)

            # Add the canvas to the layout
            self.graph_layout.addWidget(self.graph_canvas)

        except Exception as e:
            QMessageBox.critical(self.parent, "Error", f"An error occurred while visualizing data: {str(e)}")

    def display_excel_contents(self, excel_data):
        """Display Excel data with buttons on the right and additional components below."""
        dialog = QDialog(self.parent)
        dialog.setWindowTitle("Excel File Contents")
        dialog.setStyleSheet("background-color: #cce7ff;")
        dialog.resize(1200, 1000)

        # Center the dialog on the screen
        screen = QApplication.primaryScreen().geometry()
        dialog_geometry = dialog.frameGeometry()
        dialog_geometry.moveCenter(screen.center())
        dialog.move(dialog_geometry.topLeft())

        # Main vertical layout
        main_layout = QVBoxLayout()

        # Horizontal layout for table and right-side buttons
        top_horizontal_layout = QHBoxLayout()


        # Tab widget for displaying Excel sheets
        self.tab_widget = QTabWidget()  # Assign to self.tab_widget
        self.tab_widget.setStyleSheet("""
        QTabWidget::pane {
            border: 1px solid #d4d4d4;
        }
        QTabBar::tab {
            padding: 10px 20px;  /* Increase tab padding */
            font-size: 14px;
            background-color: #e6f2ff;
            border: 1px solid #d4d4d4;
            margin-right: 5px; /* Add spacing between tabs */
        }
        QTabBar::tab:selected {
            background-color: #4CAF50;
            color: white;
            border: 1px solid #4CAF50;
        }
    """)

        # Store sheet data in a dictionary to track by tab
        self.sheet_dict = {}  # Store all sheets
        for sheet_name, sheet_data in excel_data.items():
            sheet_data.columns = sheet_data.columns.str.lower()
            sheet_data = sheet_data.fillna("")

        # Table widget for sheet
            table_widget = QTableWidget()
            table_widget.setRowCount(sheet_data.shape[0])
            table_widget.setColumnCount(sheet_data.shape[1])
            table_widget.setHorizontalHeaderLabels(sheet_data.columns)

        # Enable scrollbars
        table_widget.setHorizontalScrollBarPolicy(Qt.ScrollBarAsNeeded)
        table_widget.setVerticalScrollBarPolicy(Qt.ScrollBarAsNeeded)

        # Adjust column widths to prevent clipping
        for col_index in range(sheet_data.shape[1]):
            table_widget.setColumnWidth(col_index, 200)

        # Allow resizable columns
        table_widget.horizontalHeader().setSectionResizeMode(QHeaderView.Interactive)

        # Set column widths explicitly
        for col_index in range(sheet_data.shape[1]):
            table_widget.setColumnWidth(col_index, 200)

        # Populate table data
            for i in range(sheet_data.shape[0]):
                for j in range(sheet_data.shape[1]):
                    item = QTableWidgetItem(str(sheet_data.iat[i, j]))
                    item.setTextAlignment(Qt.AlignRight | Qt.AlignVCenter)
                    table_widget.setItem(i, j, item)

            table_widget.horizontalHeader().setSectionResizeMode(QHeaderView.Stretch)

            # Add sheet data to the dictionary
            self.sheet_dict[sheet_name] = sheet_data

            # Add the table widget to the tab widget
            self.tab_widget.addTab(table_widget, sheet_name)

        # Ensure self.sheet_data is set to the first sheet by default
        def update_current_sheet(index):
            selected_sheet_name = self.tab_widget.tabText(index)
            self.sheet_data = self.sheet_dict[selected_sheet_name]
            print(f"Current Sheet: {selected_sheet_name}")

        self.tab_widget.currentChanged.connect(update_current_sheet)
        if self.tab_widget.count() > 0:
            update_current_sheet(0)

        # Add the tab widget to the horizontal layout
        top_horizontal_layout.addWidget(self.tab_widget)


        # Right-side layout for buttons and graph
        right_button_layout = QVBoxLayout()

        # Reduce spacing between buttons and margins
        right_button_layout.setSpacing(5)  # Set spacing between widgets to 5 pixels
        right_button_layout.setContentsMargins(0, 0, 0, 0)  # Remove margins around the layout

        # Add button to download sheets into a new Excel file
        download_button = QPushButton("Download Sheets as Excel")
        download_button.setStyleSheet("font-size: 16px; color: white; background-color: #4CAF50;")
        download_button.clicked.connect(self.download_sheets_as_excel)
        right_button_layout.addWidget(download_button)

        # Add Data Button
        add_data_button = QPushButton("Add Data")
        add_data_button.setStyleSheet("font-size: 16px; color: white; background-color: #4CAF50;")
        add_data_button.setFixedHeight(35)  # Reduce button height
        add_data_button.clicked.connect(self.add_data_to_sheet)
        right_button_layout.addWidget(add_data_button)

        # Remove Data Button
        remove_data_button = QPushButton("Remove Data")
        remove_data_button.setStyleSheet("font-size: 16px; color: white; background-color: #F44336;")
        remove_data_button.setFixedHeight(35)
        remove_data_button.clicked.connect(self.remove_data_from_sheet)
        right_button_layout.addWidget(remove_data_button)

        # Visualize Data Button
        visualize_button = QPushButton("Visualize Data")
        visualize_button.setStyleSheet("font-size: 16px; color: white; background-color: #4CAF50;")
        visualize_button.setFixedHeight(35)
        visualize_button.clicked.connect(self.visualize_data)
        right_button_layout.addWidget(visualize_button)

        # Grouping Buttons
        group_month_button = QPushButton("Group by Month")
        group_month_button.setStyleSheet("font-size: 16px;")
        group_month_button.setFixedHeight(35)
        group_month_button.clicked.connect(self.group_by_month)
        right_button_layout.addWidget(group_month_button)

        group_fund_button = QPushButton("Group by Fund Number")
        group_fund_button.setStyleSheet("font-size: 16px;")
        group_fund_button.setFixedHeight(35)
        group_fund_button.clicked.connect(self.group_by_fund)
        right_button_layout.addWidget(group_fund_button)

        # Sum Costs by Month Button
        sum_by_month_button = QPushButton("Sum Costs by Month")
        sum_by_month_button.setStyleSheet("font-size: 16px;")
        sum_by_month_button.setFixedHeight(35)
        sum_by_month_button.clicked.connect(self.sum_costs_by_month)
        right_button_layout.addWidget(sum_by_month_button)

        # Sum Costs by Fund Button
        sum_by_fund_button = QPushButton("Sum Costs by Fund")
        sum_by_fund_button.setStyleSheet("font-size: 16px;")
        sum_by_fund_button.setFixedHeight(35)
        sum_by_fund_button.clicked.connect(self.sum_costs_by_fund)
        right_button_layout.addWidget(sum_by_fund_button)

        # Categorize Items Button
        categorize_button = QPushButton("Categorize Items")
        categorize_button.setStyleSheet("font-size: 16px; color: white; background-color: #4CAF50;")
        categorize_button.setFixedHeight(35)
        categorize_button.clicked.connect(self.categorize_items)
        right_button_layout.addWidget(categorize_button)

        # Categorize and Group Items Button
        categorize_button = QPushButton("Categorize and Group Items")
        categorize_button.setStyleSheet("font-size: 16px; color: white; background-color: #4CAF50;")
        categorize_button.setFixedHeight(35)
        categorize_button.clicked.connect(self.categorize_and_group_items)
        right_button_layout.addWidget(categorize_button)


        # Graph Layout (below buttons)
        self.graph_layout = QVBoxLayout()
        right_button_layout.addLayout(self.graph_layout)

        # Add right button layout to the horizontal layout
        top_horizontal_layout.addLayout(right_button_layout)

        # Add the horizontal layout to the main vertical layout
        main_layout.addLayout(top_horizontal_layout)

        # Bottom layout for cost, date range filter, and grant allocation
        bottom_layout = QVBoxLayout()

        # Total Cost and Selected Sum Labels
        total_cost_label = QLabel(f"Total Cost: ${self.total_cost:.2f}")
        total_cost_label.setStyleSheet("font-size: 16px; color: black;")
        bottom_layout.addWidget(total_cost_label)

        self.selected_sum_label = QLabel("Selected Sum: $0.00")
        self.selected_sum_label.setStyleSheet("font-size: 16px; color: black;")
        bottom_layout.addWidget(self.selected_sum_label)

        # Date range filtering
        date_range_layout = QHBoxLayout()

        start_date_label = QLabel("Start Date:")
        start_date_label.setStyleSheet("font-size: 16px; color: black;")
        date_range_layout.addWidget(start_date_label)

        self.start_date_edit = QDateEdit()
        self.start_date_edit.setCalendarPopup(True)
        self.start_date_edit.setDate(QDate.currentDate().addYears(-1))
        date_range_layout.addWidget(self.start_date_edit)

        end_date_label = QLabel("End Date:")
        end_date_label.setStyleSheet("font-size: 16px; color: black;")
        date_range_layout.addWidget(end_date_label)

        self.end_date_edit = QDateEdit()
        self.end_date_edit.setCalendarPopup(True)
        self.end_date_edit.setDate(QDate.currentDate())
        date_range_layout.addWidget(self.end_date_edit)

        filter_button = QPushButton("Filter Costs by Date Range")
        filter_button.setStyleSheet("font-size: 16px; color: white; background-color: #4CAF50;")
        filter_button.clicked.connect(self.filter_costs_by_date)
        date_range_layout.addWidget(filter_button)

        bottom_layout.addLayout(date_range_layout)

        # Grant allocation
        grant_layout = QHBoxLayout()

        grant_label = QLabel("Select Grant:")
        grant_label.setStyleSheet("font-size: 16px; color: black;")
        grant_layout.addWidget(grant_label)

        self.grant_combo = QComboBox()
        self.grant_combo.addItems(self.grant_management.get_grant_names())
        grant_layout.addWidget(self.grant_combo)

        allocate_button = QPushButton("Allocate Selected Costs to Grant")
        allocate_button.setStyleSheet("font-size: 16px; color: white; background-color: #4CAF50;")
        allocate_button.clicked.connect(self.allocate_costs_to_grant)
        grant_layout.addWidget(allocate_button)

        net_amount_label = QLabel("Net Amount in Grant: $0.00")
        net_amount_label.setStyleSheet("font-size: 16px; color: black;")
        self.net_amount_label = net_amount_label
        grant_layout.addWidget(net_amount_label)

        bottom_layout.addLayout(grant_layout)

        # Add bottom layout to the main vertical layout
        main_layout.addLayout(bottom_layout)

        # Set the dialog layout
        dialog.setLayout(main_layout)
        dialog.exec_()

    def create_table_widget(self, sheet_data):
        """Helper function to create a QTableWidget from sheet data."""
        table_widget = QTableWidget()
        table_widget.setRowCount(sheet_data.shape[0])
        table_widget.setColumnCount(sheet_data.shape[1])
        table_widget.setHorizontalHeaderLabels(sheet_data.columns)
        for i in range(sheet_data.shape[0]):
            for j in range(sheet_data.shape[1]):
                item = QTableWidgetItem(str(sheet_data.iat[i, j]))
                item.setTextAlignment(Qt.AlignRight | Qt.AlignVCenter)
                table_widget.setItem(i, j, item)
        table_widget.horizontalHeader().setSectionResizeMode(QHeaderView.Stretch)
        return table_widget

    def download_sheets_as_excel(self):
        """Allow saving all sheets as a new Excel file."""
        try:
            options = QFileDialog.Options()
            file_path, _ = QFileDialog.getSaveFileName(
                self.parent, "Save Sheets As Excel", "output_sheets.xlsx", "Excel Files (*.xlsx);;All Files (*)", options=options
            )

            if file_path:
                # Save all sheets in the current sheet dictionary
                with pd.ExcelWriter(file_path, engine='xlsxwriter') as writer:
                    for sheet_name, sheet_data in self.sheet_dict.items():
                        sheet_data.to_excel(writer, sheet_name=sheet_name, index=False)

                QMessageBox.information(self.parent, "Success", f"All sheets have been saved to:\n{file_path}")
        except Exception as e:
            QMessageBox.critical(self.parent, "Error", f"An error occurred while saving the file:\n{str(e)}")        

    def display_grouped_data_with_highlights(self, grouped_data, title, column_name):
        """
        Display grouped data with row highlights and ensure scrollbars and content are visible.
        """
        dialog = QDialog(self.parent)
        dialog.setWindowTitle(title)
        dialog.resize(1000, 600)  # Set larger initial size

        layout = QVBoxLayout()

        # Table widget to display grouped data
        table_widget = QTableWidget()
        table_widget.setRowCount(grouped_data.shape[0])
        table_widget.setColumnCount(grouped_data.shape[1])
        table_widget.setHorizontalHeaderLabels(grouped_data.columns)

        # Function to generate a random color
        def generate_random_color():
            return QColor(
                random.randint(100, 255),  # Red (avoid too dark)
                random.randint(100, 255),  # Green
                random.randint(100, 255),  # Blue
                200  # Alpha for slight transparency
            )

        # Ensure each unique group gets a unique color
        color_mapping = {}
        used_colors = set()

        for i in range(grouped_data.shape[0]):
            group_value = grouped_data.iloc[i][column_name]

            # Assign a unique color if not already assigned
            if group_value not in color_mapping:
                while True:
                    new_color = generate_random_color()
                    color_tuple = (new_color.red(), new_color.green(), new_color.blue())
                    if color_tuple not in used_colors:  # Ensure uniqueness
                        used_colors.add(color_tuple)
                        color_mapping[group_value] = new_color
                        break

            row_color = color_mapping[group_value]
            for j in range(grouped_data.shape[1]):
                item = QTableWidgetItem(str(grouped_data.iat[i, j]))
                item.setTextAlignment(Qt.AlignCenter)
                item.setBackground(row_color)
                table_widget.setItem(i, j, item)

        table_widget.horizontalHeader().setSectionResizeMode(QHeaderView.Stretch)
        layout.addWidget(table_widget)


        # Enable scrollbars
        table_widget.setHorizontalScrollBarPolicy(Qt.ScrollBarAsNeeded)
        table_widget.setVerticalScrollBarPolicy(Qt.ScrollBarAsNeeded)

        # Allow user to resize columns
        table_widget.horizontalHeader().setSectionResizeMode(QHeaderView.Interactive)

        # Set column widths explicitly (adjust as needed)
        for col_index in range(grouped_data.shape[1]):
            table_widget.setColumnWidth(col_index, 200)


        # Highlight rows by unique values in the specified column
        for i in range(grouped_data.shape[0]):
            group_value = grouped_data.iloc[i][column_name]
            if group_value not in color_mapping:
                color_mapping[group_value] = colors[color_index % len(colors)]
                color_index += 1

            row_color = color_mapping[group_value]
            for j in range(grouped_data.shape[1]):
                item = QTableWidgetItem(str(grouped_data.iat[i, j]))
                item.setTextAlignment(Qt.AlignCenter)
                item.setBackground(row_color)
                table_widget.setItem(i, j, item)

        layout.addWidget(table_widget)

        # Buttons
        save_button = QPushButton("Save Grouped Data")
        save_button.clicked.connect(lambda: self.save_grouped_data_with_highlights(grouped_data, title, column_name))
        layout.addWidget(save_button)

        add_sheet_button = QPushButton("Add Grouped Data as New Sheet")
        add_sheet_button.clicked.connect(lambda: self.add_grouped_data_as_new_sheet(grouped_data, title))
        layout.addWidget(add_sheet_button)

        dialog.setLayout(layout)
        dialog.exec_()

    def generate_color_mapping(self, grouped_data, column_name):
        """
        Generate or update a color mapping for unique values in the specified column.
        Colors are stored as hex strings for compatibility with both PyQt and Excel.
        This ensures consistent coloring across grouped data and saved Excel files.
        """
        # Ensure group_color_mapping exists to store group-to-color associations
        if not hasattr(self, 'group_color_mapping'):
            self.group_color_mapping = {}

        # Track already used colors to avoid duplicates
        used_colors = set(self.group_color_mapping.values())

        def generate_random_hex_color():
            """
            Generate a random hex color code in the format '#RRGGBB'.
            The color components (R, G, B) are randomly chosen to ensure
            bright and distinguishable colors.
            """
            return "#{:02X}{:02X}{:02X}".format(
                random.randint(100, 255),  # Red component
                random.randint(100, 255),  # Green component
                random.randint(100, 255)   # Blue component
            )

        # Iterate through unique group values in the specified column
        for group_value in grouped_data[column_name].unique():
            # If the group does not have an assigned color, generate one
            if group_value not in self.group_color_mapping:
                while True:
                    # Generate a new random hex color
                    new_color = generate_random_hex_color()

                    # Ensure the color is not already in use
                    if new_color not in used_colors:
                        self.group_color_mapping[group_value] = new_color  # Store the color for the group
                        used_colors.add(new_color)  # Add to the set of used colors
                        break  # Exit the loop once a unique color is found





    def display_grouped_data_with_repeats(self, grouped_data, title):
        """
        Display grouped data with repeats in a popup.
        Add buttons to save grouped data to a file or as a new sheet.
        Persist random colors for each group.
        """
        if self.sheet_data is None:
            QMessageBox.warning(self.parent, "No Data", "No data has been loaded. Please upload an Excel file first.")
            return

        # Dialog setup
        dialog = QDialog(self.parent)
        dialog.setWindowTitle(title)
        dialog.resize(800, 600)

        layout = QVBoxLayout()

        # Table widget to display grouped data
        table_widget = QTableWidget()
        table_widget.setRowCount(grouped_data.shape[0])
        table_widget.setColumnCount(grouped_data.shape[1])
        table_widget.setHorizontalHeaderLabels([str(col).capitalize() for col in grouped_data.columns])

        # Determine grouping column dynamically
        column_name = "month" if "month" in grouped_data.columns else "fund_number"

        # Generate or update group colors
        self.generate_color_mapping(grouped_data, column_name)

        # Populate the table with grouped data and apply colors
        for i in range(grouped_data.shape[0]):
            group_value = grouped_data.iloc[i][column_name]
            row_color = self.group_color_mapping[group_value]

            for j in range(grouped_data.shape[1]):
                item = QTableWidgetItem(str(grouped_data.iat[i, j]))
                item.setTextAlignment(Qt.AlignCenter)
                item.setBackground(row_color)
                table_widget.setItem(i, j, item)

        table_widget.horizontalHeader().setSectionResizeMode(QHeaderView.Stretch)
        layout.addWidget(table_widget)

        # Buttons layout at the bottom
        button_layout = QHBoxLayout()

        # Save Grouped Data Button
        save_button = QPushButton("Save Grouped Data")
        save_button.setStyleSheet("""
            QPushButton {
                font-size: 16px;
                color: white;
                background-color: #4CAF50;
                padding: 10px;
                border-radius: 5px;
            }
            QPushButton:hover {
                background-color: #45a049;
            }
        """)
        save_button.clicked.connect(lambda: self.save_grouped_data_with_highlights(grouped_data, title, column_name))
        button_layout.addWidget(save_button)

        # Add Grouped Data as New Sheet Button
        add_sheet_button = QPushButton("Add Grouped Data as New Sheet")
        add_sheet_button.setStyleSheet("""
            QPushButton {
                font-size: 16px;
                color: white;
                background-color: #2196F3;
                padding: 10px;
                border-radius: 5px;
            }
            QPushButton:hover {
                background-color: #1976D2;
            }
        """)
        add_sheet_button.clicked.connect(lambda: self.add_grouped_data_as_new_sheet(grouped_data, title))
        button_layout.addWidget(add_sheet_button)

        layout.addLayout(button_layout)

        dialog.setLayout(layout)
        dialog.exec_()



    def add_grouped_data_as_new_sheet(self, grouped_data, title, save_to_file=False):
        """
        Add the grouped data as a new sheet in the main Excel display, update the tabs,
        and optionally save the grouped data into an Excel file.
        Handles sheet name length restrictions and row highlighting.
        """
        try:
            # Step 1: Truncate the title to meet the 31-character limit
            base_name = title.replace(' ', '_').lower()[:27]  # Reserve space for suffixes like _1, _2, etc.
            new_sheet_name = base_name

            # Step 2: Ensure the name is unique
            counter = 1
            while new_sheet_name in self.sheet_dict:
                new_sheet_name = f"{base_name}_{counter}"
                counter += 1

            # Step 3: Add the grouped data to the sheet dictionary
            self.sheet_dict[new_sheet_name] = grouped_data
            print(f"Added new sheet: {new_sheet_name}")  # Debugging statement

            # Step 4: Create a new QTableWidget for the grouped data
            new_table_widget = QTableWidget()
            new_table_widget.setRowCount(grouped_data.shape[0])
            new_table_widget.setColumnCount(grouped_data.shape[1])
            new_table_widget.setHorizontalHeaderLabels([str(col).capitalize() for col in grouped_data.columns])

            # Highlight rows for grouped data
            color_palette = [
                QColor("#FFCCCC"), QColor("#CCFFCC"), QColor("#CCCCFF"),
                QColor("#FFFF99"), QColor("#FFCCFF"), QColor("#99FFFF")
            ]
            color_mapping = {}
            color_index = 0

            for i in range(grouped_data.shape[0]):
                group_key = grouped_data.iloc[i]['fund_number'] if 'fund_number' in grouped_data.columns else grouped_data.iloc[i]['month']
                if group_key not in color_mapping:
                    color_mapping[group_key] = color_palette[color_index % len(color_palette)]
                    color_index += 1

                row_color = color_mapping[group_key]
                for j in range(grouped_data.shape[1]):
                    item = QTableWidgetItem(str(grouped_data.iat[i, j]))
                    item.setTextAlignment(Qt.AlignRight | Qt.AlignVCenter)
                    item.setBackground(row_color)
                    new_table_widget.setItem(i, j, item)

            new_table_widget.horizontalHeader().setSectionResizeMode(QHeaderView.Stretch)

            # Step 5: Add the new table widget as a tab
            if not hasattr(self, 'tab_widget') or self.tab_widget is None:
                QMessageBox.warning(self.parent, "Error", "Tab widget is not initialized. Please upload data first.")
                return

            self.tab_widget.addTab(new_table_widget, new_sheet_name)
            self.tab_widget.setCurrentWidget(new_table_widget)  # Switch focus to the new tab

            # Update the current sheet data
            self.sheet_data = grouped_data

            # Step 6: Save to file if requested
            if save_to_file:
                save_path, _ = QFileDialog.getSaveFileName(
                    self.parent,
                    "Save Grouped Data",
                    f"{new_sheet_name}.xlsx",
                    "Excel Files (*.xlsx);;All Files (*)"
                )
                if save_path:
                    grouped_data.to_excel(save_path, index=False)
                    QMessageBox.information(self.parent, "Success", f"Grouped data saved to: {save_path}")

            # Success message
            QMessageBox.information(self.parent, "Success", f"Grouped data has been added as a new sheet: '{new_sheet_name}'.")

        except Exception as e:
            QMessageBox.critical(self.parent, "Error", f"An error occurred while adding the new sheet:\n{str(e)}")



    def group_by_month(self):
        """Group data by Month and display with row highlights."""
        if self.sheet_data is None or 'expiration date' not in self.sheet_data.columns:
            QMessageBox.warning(self.parent, "Error", "'Expiration Date' column is missing.")
            return

        try:
            # Convert 'expiration date' to datetime and extract the month
            self.sheet_data['expiration date'] = pd.to_datetime(self.sheet_data['expiration date'], errors='coerce')
            self.sheet_data['month'] = self.sheet_data['expiration date'].dt.to_period('M').astype(str)

            # Sort the data
            grouped_data = self.sheet_data.sort_values(by='month')

            # Generate colors for unique groups
            self.generate_color_mapping(grouped_data, 'month')
            self.display_grouped_data_with_highlights(grouped_data, "Grouped by Month", 'month')
        except Exception as e:
            QMessageBox.critical(self.parent, "Error", f"An error occurred during grouping by month: {str(e)}")


    def group_by_fund(self):
        """Group data by Fund Number and display with row highlights."""
        if self.sheet_data is None or 'fund_number' not in self.sheet_data.columns:
            QMessageBox.warning(self.parent, "Error", "'Fund Number' column is missing.")
            return

        try:
            # Group and display the data
            grouped_data = self.sheet_data.sort_values(by='fund_number')
            self.display_grouped_data_with_highlights(grouped_data, "Grouped by Fund Number", 'fund_number')
            # Color Mapping
            # Generate colors for unique groups
            self.generate_color_mapping(grouped_data, 'fund_number')
            self.display_grouped_data_with_highlights(grouped_data, "Grouped by Fund Number", 'fund_number')
        except Exception as e:
            QMessageBox.critical(self.parent, "Error", f"An error occurred during grouping by fund: {str(e)}")
        

    def group_by_column(self, column_name, title):
        """
        Generic method to group data by a column, display in a popup, and highlight rows.
        """
        if self.sheet_data is None or column_name not in self.sheet_data.columns:
            QMessageBox.warning(self.parent, "Error", f"'{column_name}' column is missing in the data.")
            return

        try:
            # Group and sort data
            grouped_data = self.sheet_data.sort_values(by=column_name)
            # Color mapping
            self.group_color_mapping = {}
            
            # Create and display a table with highlighted rows
            self.display_grouped_data_with_highlights(grouped_data, title, column_name)
        except Exception as e:
            QMessageBox.critical(self.parent, "Error", f"An error occurred during grouping: {str(e)}")


    def sum_costs_by_month(self):
        """Sum the costs for each individual month and display in a popup."""
        if self.sheet_data is None or 'month' not in self.sheet_data.columns or 'cost' not in self.sheet_data.columns:
            QMessageBox.warning(self.parent, "No Grouped Data", "Please group data by Month first and ensure 'Cost' column exists.")
            return

        try:
            # Clean and convert cost column to numeric
            self.sheet_data['clean_cost'] = self.sheet_data['cost'].apply(lambda x: self.clean_and_convert_cost(x))

            # Sum costs by month
            summed_data = self.sheet_data.groupby('month')['clean_cost'].sum().reset_index()
            summed_data.rename(columns={'clean_cost': 'total_cost'}, inplace=True)

            # Use existing color mapping for months
            self.display_summarized_data_with_colors(summed_data, "Summed Costs by Month", self.group_color_mapping)
        except Exception as e:
            QMessageBox.critical(self.parent, "Error", f"An error occurred while summing costs by month: {str(e)}")

    def sum_costs_by_fund(self):
        """Sum the costs for each individual fund and display in a popup."""
        if self.sheet_data is None or 'fund_number' not in self.sheet_data.columns or 'cost' not in self.sheet_data.columns:
            QMessageBox.warning(self.parent, "No Grouped Data", "Please group data by Fund first and ensure 'Cost' column exists.")
            return

        try:
            # Clean and convert cost column to numeric
            self.sheet_data['clean_cost'] = self.sheet_data['cost'].apply(lambda x: self.clean_and_convert_cost(x))

            # Sum costs by fund
            summed_data = self.sheet_data.groupby('fund_number')['clean_cost'].sum().reset_index()
            summed_data.rename(columns={'clean_cost': 'total_cost'}, inplace=True)

            # Use existing color mapping for fund numbers
            self.display_summarized_data_with_colors(summed_data, "Summed Costs by Fund", self.group_color_mapping)
        except Exception as e:
            QMessageBox.critical(self.parent, "Error", f"An error occurred while summing costs by fund: {str(e)}")
        
    def display_summarized_data_popup(self, summarized_data, title, color_mapping):
        """
        Display summarized data in a popup with an option to add it as a new sheet, retaining original group colors.
        Fixes missing colors and ensures data is visible in table cells.
        """
        dialog = QDialog(self.parent)
        dialog.setWindowTitle(title)
        dialog.resize(800, 600)

        layout = QVBoxLayout()
        table_widget = QTableWidget()
        table_widget.setRowCount(summarized_data.shape[0])
        table_widget.setColumnCount(summarized_data.shape[1])
        table_widget.setHorizontalHeaderLabels([str(col).capitalize() for col in summarized_data.columns])

        # Populate the table and apply colors
        for i in range(summarized_data.shape[0]):
            group_value = summarized_data.iloc[i, 0]  # First column (group key)
            row_color = color_mapping.get(group_value, QColor("#FFFFFF"))  # Default white if no color found

            for j in range(summarized_data.shape[1]):
                item_text = str(summarized_data.iat[i, j])
                item = QTableWidgetItem(item_text)
                item.setTextAlignment(Qt.AlignCenter)  # Align text properly
                item.setBackground(row_color)  # Apply background color to the cell
                table_widget.setItem(i, j, item)

        table_widget.horizontalHeader().setSectionResizeMode(QHeaderView.Stretch)
        layout.addWidget(table_widget)

        # Add "Add as New Sheet" Button
        add_sheet_button = QPushButton("Add as New Sheet")
        add_sheet_button.setStyleSheet("""
            QPushButton {
                font-size: 16px;
                background-color: #4CAF50;
                color: white;
                padding: 10px;
                border-radius: 5px;
            }
            QPushButton:hover {
                background-color: #45a049;
            }
        """)
        add_sheet_button.clicked.connect(lambda: self.add_summarized_data_as_sheet(summarized_data, title, color_mapping))
        layout.addWidget(add_sheet_button)

        dialog.setLayout(layout)
        dialog.exec_()



    def sum_costs_by_month(self):
        """Sum the costs for each month and display with colors matching the grouped data."""
        if self.sheet_data is None or 'month' not in self.sheet_data.columns or 'cost' not in self.sheet_data.columns:
            QMessageBox.warning(self.parent, "No Grouped Data", "Please group data by Month first and ensure 'Cost' column exists.")
            return

        try:
            # Clean and convert cost column to numeric
            self.sheet_data['clean_cost'] = self.sheet_data['cost'].apply(self.clean_and_convert_cost)

            # Sum costs by month
            summed_data = self.sheet_data.groupby('month')['clean_cost'].sum().reset_index()
            summed_data.rename(columns={'clean_cost': 'total_cost'}, inplace=True)

            # Use the existing color mapping for months
            self.display_summarized_data_with_colors(summed_data, "Summed Costs by Month", self.group_color_mapping)

        except Exception as e:
            QMessageBox.critical(self.parent, "Error", f"An error occurred while summing costs by month: {str(e)}")


    def sum_costs_by_fund(self):
        """Sum the costs for each fund and display with colors matching the grouped data."""
        if self.sheet_data is None or 'fund_number' not in self.sheet_data.columns or 'cost' not in self.sheet_data.columns:
            QMessageBox.warning(self.parent, "No Grouped Data", "Please group data by Fund first and ensure 'Cost' column exists.")
            return

        try:
            # Clean and convert cost column to numeric
            self.sheet_data['clean_cost'] = self.sheet_data['cost'].apply(self.clean_and_convert_cost)

            # Sum costs by fund
            summed_data = self.sheet_data.groupby('fund_number')['clean_cost'].sum().reset_index()
            summed_data.rename(columns={'clean_cost': 'total_cost'}, inplace=True)

            # Use the existing color mapping for fund numbers
            self.display_summarized_data_with_colors(summed_data, "Summed Costs by Fund", self.group_color_mapping)

        except Exception as e:
            QMessageBox.critical(self.parent, "Error", f"An error occurred while summing costs by fund: {str(e)}")


    def display_summarized_data_with_colors(self, summarized_data, title, color_mapping):
        """
        Display summarized data with colors consistent with original grouped data.
        """
        dialog = QDialog(self.parent)
        dialog.setWindowTitle(title)
        dialog.resize(1000, 800)

        layout = QVBoxLayout()

        # Table widget to display summarized data
        table_widget = QTableWidget()
        table_widget.setRowCount(summarized_data.shape[0])
        table_widget.setColumnCount(summarized_data.shape[1])
        table_widget.setHorizontalHeaderLabels(summarized_data.columns)

        # Populate the table and apply colors
        for i in range(summarized_data.shape[0]):
            group_value = summarized_data.iloc[i, 0]  # Group key (e.g., month or fund_number)

            # Retrieve color from the mapping, default to white
            row_color = color_mapping.get(group_value, (255, 255, 255, 255))
            qcolor = QColor(row_color)  # Convert tuple to QColor

            for j in range(summarized_data.shape[1]):
                item = QTableWidgetItem(str(summarized_data.iat[i, j]))
                item.setTextAlignment(Qt.AlignCenter)
                item.setBackground(qcolor)  # Apply background color
                table_widget.setItem(i, j, item)

        table_widget.horizontalHeader().setSectionResizeMode(QHeaderView.Stretch)
        layout.addWidget(table_widget)

        # Save as New Sheet button
        save_button = QPushButton("Save as New Sheet")
        save_button.clicked.connect(lambda: self.add_summarized_data_as_sheet(summarized_data, title, color_mapping))
        layout.addWidget(save_button)

        dialog.setLayout(layout)
        dialog.exec_()



    def add_summarized_data_as_sheet(self, summarized_data, title, color_mapping):
        """
        Add summarized data as a new sheet with retained group colors.
        """
        sheet_name = title.replace(" ", "_").lower()
        counter = 1
        while sheet_name in self.sheet_dict:
            sheet_name = f"{title.replace(' ', '_').lower()}_{counter}"
            counter += 1

        # Store the summarized data in the sheet dictionary
        self.sheet_dict[sheet_name] = summarized_data

        # Create a QTableWidget for the new sheet
        table_widget = QTableWidget()
        table_widget.setRowCount(summarized_data.shape[0])
        table_widget.setColumnCount(summarized_data.shape[1])
        table_widget.setHorizontalHeaderLabels(summarized_data.columns)

        # Apply group colors
        for i in range(summarized_data.shape[0]):
            group_value = summarized_data.iloc[i, 0]
            row_color = self.group_color_mapping.get(group_value, (255, 255, 255, 255))  # Default to white
            qcolor = QColor(row_color)  # Reconstruct QColor from the tuple

            for j in range(summarized_data.shape[1]):
                item = QTableWidgetItem(str(summarized_data.iat[i, j]))  # Create the QTableWidgetItem
                item.setTextAlignment(Qt.AlignCenter)
                item.setBackground(qcolor)  # Set background color
                table_widget.setItem(i, j, item)

        table_widget.horizontalHeader().setSectionResizeMode(QHeaderView.Stretch)
        self.tab_widget.addTab(table_widget, sheet_name)
        self.tab_widget.setCurrentWidget(table_widget)

        QMessageBox.information(self.parent, "Success", f"Summarized data added as a new sheet: {sheet_name}")



    def clean_and_convert_cost(self, value):
        """
        Clean and convert a cost value to a numeric format.
        Supports values like '$3.30' and '3.30'.
        """
        try:
            if isinstance(value, str):
                # Remove any non-numeric characters except '.' and convert to float
                return float(re.sub(r'[^\d.]', '', value))
            elif isinstance(value, (int, float)):
                return float(value)  # Already a valid number
            else:
                return 0.0  # Default for invalid values
        except Exception:
            return 0.0  # Default for errors during conversion



    def update_selected_sum(self, table_widget):
        """Update the sum of selected costs based on highlighted rows, specifically from the 'cost' column."""
        selected_sum = 0.0
        if isinstance(table_widget, QTableWidget):
            # Find the index of the 'cost' column
            headers = [table_widget.horizontalHeaderItem(i).text().lower() for i in range(table_widget.columnCount())]
            cost_column_index = headers.index('cost') if 'cost' in headers else -1

            if cost_column_index != -1:
                # Sum the costs only from the 'cost' column of the selected rows
                selected_rows = set(item.row() for item in table_widget.selectedItems())  # Get unique selected rows

                for row in selected_rows:
                    cost_item = table_widget.item(row, cost_column_index)
                    if cost_item:
                        try:
                            cost_value = float(re.sub(r'[^\d.]', '', cost_item.text()))
                            selected_sum += cost_value
                        except ValueError:
                            continue  # Skip non-numeric values

        self.selected_sum_label.setText(f"Selected Sum: ${selected_sum:.2f}")

    def filter_costs_by_date(self):
        """Filter costs based on the selected date range."""
        if self.sheet_data is None or 'expiration date' not in self.sheet_data.columns:
            QMessageBox.warning(self.parent, "No Date Data", "No 'Expiration Date' column found for filtering.")
            return

        start_date = self.start_date_edit.date().toPyDate()
        end_date = self.end_date_edit.date().toPyDate()

        filtered_data = self.sheet_data[
            (self.sheet_data['expiration date'] >= pd.to_datetime(start_date)) &
            (self.sheet_data['expiration date'] <= pd.to_datetime(end_date))
        ]

        if 'cost' in filtered_data.columns:
            cleaned_costs = filtered_data['cost'].apply(lambda x: float(re.sub(r'[^\d.]', '', str(x))) if x != '' else 0)
            total_filtered_cost = cleaned_costs.sum()
            QMessageBox.information(self.parent, "Filtered Costs", f"Total Costs in Date Range: ${total_filtered_cost:.2f}")
        else:
            QMessageBox.warning(self.parent, "Cost Column Missing", "'Cost' column not found in the filtered data.")



### replaced w/ find_name_column -> the description was the name
    def find_description_column(self):
        """Find the most likely column containing descriptions."""
        possible_names = ["description", "item", "details", "name", "product"]
        for col in self.sheet_data.columns:
            if any(keyword in col.lower() for keyword in possible_names):
                return col
        return None

    def find_name_column(self):
        """Find the most likely column containing item names."""
        possible_names = ["name", "item", "product", "details", "description"]
        for col in self.sheet_data.columns:
            if any(keyword in col.lower() for keyword in possible_names):
                return col
        return None

    def categorize_items(self):
        """Categorize items, group by category, and create a new sheet."""
        if self.sheet_data is None:
            QMessageBox.warning(self.parent, "No Data", "No data has been loaded. Please upload an Excel file first.")
            return

        # Extended categories with keywords
        categories = {
            'Media': [
                # General media
                'media', 'PBS', 'cell-culture', 'cell culture media', 'DMEM', 'RPMI', 'EMEM', 
                'McCoy', 'IMDM', 'F-12', 'Ham\'s F-12', 'MEM', 'AMEM', 'α-MEM', 'Basal Medium Eagle',
                'L-15', 'Leibovitz\'s L-15', 'Hank\'s Balanced Salt Solution', 'HBSS', 'Eagle\'s Medium', 
                'Williams\' Medium E', 'Coon\'s Modified Ham\'s F-12', 'serum-free medium', 
                'low-glucose medium', 'high-glucose medium', 'DMEM/F-12', 'RPMI-1640', 'keratinocyte medium',
                
                # Neutralizers and derivatives
                'trypsin', 'trypsin-EDTA', 'trypsin neutralizer', 'TrypLE', 'trypsin substitute', 
                'neutralizing solution', 'EDTA', 'collagenase', 'dispase', 'accutase', 
                'cell dissociation solution', 'cell detachment solution', 'trypsin inhibitor',
                
                # Supplements and additives
                'glutamine', 'L-glutamine', 'sodium pyruvate', 'non-essential amino acids', 
                'NEAA', 'FBS', 'fetal bovine serum', 'bovine serum', 'horse serum', 'cell culture grade water',
                'water for injection', 'sterile water', 'distilled water', 'di water', 'ultrapure water',
                
                # Specialized media
                'neural stem cell medium', 'mesenchymal stem cell medium', 'embryonic stem cell medium',
                'organoid culture media', 'hepatocyte media', 'airway epithelial cell media',
                'fibroblast growth medium', 'skeletal muscle cell media', 'chondrocyte media', 
                'endothelial growth medium', 'epithelial cell growth medium', 'keratinocyte serum-free medium',
                
                # Growth additives
                'growth factor supplement', 'b27 supplement', 'N2 supplement', 'bfgf', 'EGF', 'insulin', 
                'transferrin', 'selenium', 'hydrocortisone', 'dexamethasone', 'ascorbic acid', 'retinoic acid'
            ],
            'W/S/N Blots': [
                'blot', 'western', 'southern', 'northern', 'gel', 'membrane', 'buffer', 'stain', 'substrate', 
                'PAGE', 'SDS-PAGE', 'acrylamide', 'electrophoresis', 'ladder', 'marker', 'staining', 'mounting',
                'HRP', 'chemiluminescent', 'chemiluminescence', 'fluorescent dye', 'immunoblot', 'immunoblotting', 
                'transfer buffer', 'running buffer', 'blotting buffer', 'wash buffer', 'blocking buffer',
                'PVDF', 'nitrocellulose', 'immobilon', 'BCA', 'Coomassie', 'silver stain', 'Ponceau', 
                'NuPAGE', 'Bis-Tris', 'Tris-Glycine', 'MES buffer', 'MOPS buffer', 'transfer membrane',
                'gel loading dye', 'protein ladder', 'DNA ladder', 'protein stain', 'anti-HRP', 'fluorescent marker',
                'secondary detection', 'imaging substrate', 'ECL', 'enhanced chemiluminescence',
                'polyacrylamide gel', 'Western substrate', 'Coomassie blue', 'chromogenic substrate',
                'hybridization buffer', 'washing reagent', 'autoradiography', 'electroblotting', 'LDS Sample Buf', 'TBS w TWEEN TBST'
            ],
            'Antibodies': [
                'antibody', 'antibodies', 'mAb', 'IgG', 'phospho', 'phospho-', 
                'rabbit', 'mouse', 'goat', 'anti-', 'affinipure', 'monoclonal',
                'secondary antibody', 'primary antibody', 'HRP-conjugated',
                'Alexa Fluor', 'AF488', 'AF568', 'AF594', 'FITC', 'APC',
                'Cy3', 'Cy5', 'Dylight', 'fluorescent antibody', 'polyclonal', 
                'isotype control', 'conjugated antibody', 'biotinylated antibody', 
                'peroxidase', 'HRP', 'AP (alkaline phosphatase)', 'ELISA antibody',
                'immunoblot antibody', 'immunohistochemistry', 'IHC', 'ICC', 
                'immunofluorescence', 'flow cytometry', 'western blot',

                ## specific antibodies
                'Hu Vimentin PE'
            ],
            'Flasks, Tips, etc.': [
                'flask', 'flasks', 'Erlenmeyer flask', 'Erlenmeyer flasks', 'Conical flask', 'Conical flasks','well', 'wells',
                'Cell culture flask', 'Cell culture flasks', 'Round-bottom flask', 'Round-bottom flasks',
                'Volumetric flask', 'Volumetric flasks', 'Vacuum flask', 'Vacuum flasks', 'Filtering flask', 'Filtering flasks',
                'tip', 'tips', 'pipet', 'pipets', 'pipette', 'pipettes', 'pipette tip', 'pipette tips',
                'filter tip', 'filter tips', 'gel-loading tip', 'gel-loading tips', 'multi-channel tip', 'multi-channel tips',
                'serological pipette', 'serological pipettes', 'manual pipette', 'manual pipettes', 'automatic pipette',
                'automatic pipettes', 'multichannel pipette', 'multichannel pipettes', 'micropipette', 'micropipettes',
                'repeater pipette', 'repeater pipettes', 'transfer pipette', 'transfer pipettes', 'glass pipette', 'glass pipettes',
                'tube', 'tubes', 'centrifuge tube', 'centrifuge tubes', 'cryogenic tube', 'cryogenic tubes', 'chambers',
                'microcentrifuge tube', 'microcentrifuge tubes', 'PCR tube', 'PCR tubes', 'glass tube', 'glass tubes',
                'Falcon tube', 'Falcon tubes', 'Eppendorf tube', 'Eppendorf tubes', 'test tube', 'test tubes',
                'storage tube', 'storage tubes', 'plts' ,'plate', 'plates', 'cell culture plate', 'cell culture plates',
                'microplate', 'microplates', 'petri plate', 'petri plates', 'ELISA plate', 'ELISA plates',
                'PCR plate', 'PCR plates', 'multi-well plate', 'multi-well plates', 'sealing plate', 'sealing plates',
                'box', 'boxes', 'storage box', 'storage boxes', 'cryogenic box', 'cryogenic boxes', 'freezer box', 'freezer boxes',
                'microtube box', 'microtube boxes', 'tip box', 'tip boxes', 'tube rack', 'tube racks',
                'autoclave-safe box', 'autoclave-safe boxes', 'syringe', 'syringes', 'disposable syringe', 'disposable syringes',
                'glass syringe', 'glass syringes', 'luer-lock syringe', 'luer-lock syringes', 'syringe filter', 'syringe filters',
                'insulin syringe', 'insulin syringes', 'rack', 'racks', 'holder', 'holders', 'pipette rack', 'pipette racks',
                'pipet rack', 'pipet racks', 'plate rack', 'plate racks', 'tube rack', 'tube racks', 'freezer rack', 'freezer racks',
                'test tube rack', 'test tube racks', 'cryovial', 'cryovials', 'cryobox', 'cryoboxes', '384', 'allprotect tissue reagent',
                'coutness', 'cryoelite', 'FBM', 'VWR BASIN'
                'nitrogen storage rack', 'sterile container', 'sterile containers', 'sample vial', 'sample vials',
                'funnel', 'funnels', 'glass slide', 'glass slides', 'coverslip', 'coverslips', 'weigh boat', 'weigh boats',
                'measuring cylinder', 'measuring cylinders', 'spray bottle', 'spray bottles', 'lab tray', 'lab trays', 'T.I.P.S.',
                'drip tray', 'drip trays', 'cell strainer', 'cell strainers', 'reservoir tray', 'reservoir trays', 'beaker', 'beakers', 'gloves', 'glove'
            ],
            'Assays': [
                'assay', 'CyQUANT', 'DNeasy', 'Glo', 'immuno', 'ChIP', 'EdU', 'FITC', 'flow cytometry', 'mycoplasma', 'purelink hipure'
            ],
            'Mouse Work': [
                'mouse', 'animal', 'rack', 'cage', 'rodent', 'bedding', 'scale', 'feeding', 'syringe for mouse', 
                'mouse holder', 'animal cage'
            ],
            'Biological': [
                # Existing items
                'Lipofectamine', 'KAPA', 'concentrator', 'concentrators', 'goat serum', 'serum', 
                'primers', 'primer', 'plasmid', 'glycerol stock', 'gBlock', 'lentivirus', 'Cas9', 'virus'
                
                # Enzymes and enzyme-related terms
                'enzyme', 'restriction enzyme', 'ligase', 'polymerase', 'reverse transcriptase',
                'DNA ligase', 'RNA polymerase', 'nuclease', 'endonuclease', 'exonuclease', 
                'DNA polymerase', 'RNase', 'RNase inhibitor', 'phosphatase', 'kinase', 
                'T4 ligase', 'Taq polymerase', 'Q5 polymerase', 'EcoRI', 'BamHI', 'NotI', 'HindIII',
                'restriction digestion', 'digestion enzyme', 'proteinase K', 'Klenow fragment',
                'DNase', 'DNAse I', 'methylase', 'NEBuilder', 'HiFi DNA Assembly', 'nickase',
                
                # NEB-specific items
                'NEB', 'New England Biolabs', 'NEBuilder HiFi', 'Q5 Master Mix', 'Quick CIP',
                'NEB ligase', 'NEB polymerase', 'NEB restriction enzyme', 'NEB buffer',
                'NEBuffer', 'NEB T4 DNA Ligase', 'NEB Taq', 'NEB EcoRI', 'NEB digestion kit',
                'NEB Phusion', 'NEB LunaScript', 'NEBNext', 'NEB methylase', 'NEB exonuclease',
                
                # Biological reagents and kits
                'competent cells', 'cloning kit', 'transfection reagent', 'DNA assembly',
                'electroporation reagent', 'viral vector', 'cDNA synthesis kit', 'PCR kit',
                'RT-PCR kit', 'NGS prep kit', 'plasmid purification', 'protein ladder', 'SuperScript', 'RNeasy',
                'marker', 'DNA ladder', 'RNA ladder', 'molecular weight marker', 'agarose', 'LB', 'agar',
                
                # Proteins and protein-related reagents
                'protein expression', 'protein purification', 'proteinase', 'protease', 
                'protein A', 'protein G', 'protein marker', 'protein standard', 'recombinant protein',
                
                # Cell culture additives and growth reagents
                'cell culture reagent', 'cell growth reagent', 'supplement', 'cell recovery medium',
                'freezing medium', 'cryopreservation', 'transfection reagent', 'nucleofection reagent',
                
                #Biologic Dyes
                'Phalloidin', 'ANNEXIN V', 'ANNEXIN', 'HOECHST', 'vimentin live cell dye', 
                'prolong diamond antifade mountant with dapi', 'dapi',

                #Ladders
                'GENERULER', 'master mix',

                # Miscellaneous biological terms
                'oligonucleotide', 'oligo', 'siRNA', 'shRNA', 'gRNA', 'sgRNA', 'RNAi', 'DNA template', 
                'RNA template', 'expression plasmid', 'vector', 'CRISPR', 'CRISPR-Cas9', 'cloning vector',
                'glycerol', 'competent cell', 'E.coli', 'BL21', 'DH5α', 'expression host', 'assembly mix', 'sequence',
                'provirus', 'glucose', 'depc-treated'


            ],
            'Drugs': [
                #Generic drugs
                'drug', 'compound', 'chemical', 'inhibitor', 'small molecule', 'antibiotic', 'penicillin', 'amoxicillin', 'ciprofloxacin', 'azithromycin', 'cephalexin', 
                'clindamycin', 'metronidazole', 'ampicillin', 'kanamycin', 'streptomycin', 'gentamicin', 'tetracycline', 'chloramphenicol', 
                'penicillin', 'carbenicillin', 'antibiotic', 'small molecule', 'compound', 'chemical', 'inhibitor', 'aspirin', 'ibuprofen', 
                'paracetamol', 'acetaminophen', 'statins', 'antiviral', 'aphidicolin', 'benzo(a)pyrene', 'doxycycline hyclate',
                
                #Antibiotics
                'penicillin', 'streptomycin', 'ampicillin', 'kanamycin', 'tetracycline', 
                'chloramphenicol', 'cephalosporin', 'erythromycin', 'rifampin', 'vancomycin', 
                'gentamicin', 'ciprofloxacin', 'levofloxacin', 'azithromycin'

                # Chemotherapy Drugs
                'cisplatin', 'carboplatin', 'oxaliplatin', 'paclitaxel', 'docetaxel', 
                'doxorubicin', 'epirubicin', 'cyclophosphamide', 'ifosfamide', 
                'etoposide', 'irinotecan', 'topotecan', 'gemcitabine', 'vincristine', 
                'vinblastine', 'vinorelbine', 'bleomycin', 'mitomycin', '5-fluorouracil', 
                'capecitabine', 'methotrexate', 'pemetrexed', 'temozolomide', 'dacarbazine', 
                'mechlorethamine', 'melphalan', 'busulfan', 'fludarabine', 'cladribine', 'OLAPARIB'],

            'Chemical': [
                'buffer', 'DMSO', 'ethanol', 'TCEP', 'methanol', 'glutaraldehyde', 
                'SDS', 'Tris', 'HEPES', 'NaCl', 'TBE', 'formaldehyde', 'molecular biology', 'molecular'
                'ammonia', 'glycine', 'crystal violet', 'ethyl cinnamate', 'iodonitrotetrazolium', 'Tetrakis(2-hydroxypropyl)',
                'poly(ethylene glycol)', 'poly-l-lysine', 'protamine sulfate grade x', 'sulfo-smcc', 'tert-butanol'
            ],
            'Services (sequencing)': ['sequencing', 'service', 'genomics', 'WGS', 'long-read', 'sequencing service'],
            'Services (one-time)': ['repair', 'installation', 'quote', 'service fee', 'one-time service', 'BSC'],
            'Services (recurrent)': ['LN2', 'nitrogen', 'maintenance', 'subscription', 'recurring service', 'FY'],
            'Office Supplies': ['ink cartridge', 'printer', 'stationery', 'WB Mason', 'pen', 'VWR Tape']
        
            }

        # Original suppliers mapped to categories
        drug_suppliers = ['MedChemExpress', 'SelleckChem', 'ApexBio']
        enzyme_suppliers = ['New England Biolabs', 'NEB']
        plasmid_suppliers = ['Addgene', 'addgene']
        antibody_suppliers = ['Cell Signaling Technology']
        office_supplies_suppliers = ['WB Mason']
        mouse_suppliers = ['SoftMouse.NET','ISEEHEAR INC']
        biological_suppliers = ['Integrated DNA Technologies', 'VectorBuilder'] 


        # Normalization mapping for suppliers
        supplier_aliases = {
            'cell signaling': 'Cell Signaling Technology',
            'sigma aldrich': 'Millipore Sigma',
            'life tech': 'Life Technologies',
            'wb mason': 'WB Mason',
            'medchemexpress': 'MedChemExpress',
            'med chem express': 'MedChemExpress',
            'medchem express': 'MedChemExpress',
            'selleckchem': 'SelleckChem',
            'selleck chem': 'SelleckChem',
            'selleckchemicals': 'SelleckChem',
            'apexbio': 'ApexBio',
            'apex bio': 'ApexBio',
            'apexbiotechnology': 'ApexBio',
            'neb': 'New England Biolabs',
            'new england biolabs': 'New England Biolabs',
            'thermo fisher': 'Thermo Fisher',
            'invitrogen': 'Invitrogen',
            'promega': 'Promega',
            'bio-rad': 'Bio-Rad',
            'qiagen': 'Qiagen',
            'takara': 'Takara',
            'roche': 'Roche',
            'clontech': 'Clontech',
            'agilent': 'Agilent',
            'millipore': 'Millipore',
            'ge healthcare': 'GE Healthcare',
            'applied biosystems': 'Applied Biosystems',
            'epicentre': 'Epicentre',
            'softmouse.net': 'SoftMouse.NET',
            'ISEEHEAR INC': 'ISEEHEAR INC',
            'ISEEHEAR': 'ISEEHEAR INC',
            'idt': 'Integrated DNA Technologies', 
            'integrated dna technologies': 'Integrated DNA Technologies',
            'integrated dna tech': 'Integrated DNA Technologies',
            'vectorbuilder': 'VectorBuilder',  # Normalize VectorBuilder
            'vector builder': 'VectorBuilder'
        }


        # Step 1: Automatically find the 'name' column
        name_column = self.find_name_column()

        # Step 2: If not found, prompt the user to select the column
        if not name_column:
            column_names = self.sheet_data.columns.tolist()
            name_column, ok = QInputDialog.getItem(
                self.parent, "Select Name Column",
                "Choose the column containing item names:", column_names, 0, False
            )
            if not ok or not name_column:
                QMessageBox.warning(self.parent, "Operation Cancelled", "No column was selected.")
                return

        # Step 3: Ensure the name column exists
        if name_column not in self.sheet_data.columns:
            QMessageBox.warning(self.parent, "Missing Column", f"The current sheet does not contain a '{name_column}' column.")
            return

        # Function to normalize supplier names
        def normalize_supplier(supplier):
            if not supplier:
                return None

            supplier_lower = supplier.lower().strip()

            # Check for known supplier patterns
            if "medchem" in supplier_lower:
                return "MedChemExpress"
            if "apex" in supplier_lower:
                return "ApexBio"
            if "selleck" in supplier_lower:
                return "SelleckChem"
            if 'neb' in supplier_lower:
                return 'New England Biolabs'
            if "idt" in supplier_lower or "integrated dna" in supplier_lower:
                return "Integrated DNA Technologies"
            if "vectorbuilder" in supplier_lower or "vector builder" in supplier_lower:
                return "VectorBuilder"

            # Return the original supplier if no match is found
            return supplier

        # Function to assign a category based on name and supplier
        def assign_category(name, supplier):
            name_lower = str(name).lower().strip()

            # Check if 'cisplatin' is in the name (prioritized)
            if "cisplatin" in name_lower:
                return "Drugs"  # Prioritize 'cisplatin'


            # Normalize supplier
            normalized_supplier = normalize_supplier(supplier)

                # Prioritized keyword matching
            if "gel ink pen" in name_lower or ("pen" in name_lower and "gel" in name_lower):
                return "Office Supplies"  # Prioritize gel ink pens
            elif "western blot" in name_lower or ("gel" in name_lower and "blot" in name_lower):
                return "W/S/N Blots"  # Secondary priority for western blot gels
            elif "gel" in name_lower:
                # Additional logic for generic gels if needed
                return "W/S/N Blots"  # Default to blot gels

            # Check supplier-based categorization
            if normalized_supplier in drug_suppliers:
                return 'Drugs'
            if normalized_supplier in enzyme_suppliers:
                return 'Biological'
            if normalized_supplier in plasmid_suppliers:
                return 'Biological'
            if normalized_supplier in antibody_suppliers:
                return 'Antibodies'
            if normalized_supplier in office_supplies_suppliers:
                return 'Office Supplies'
            # Check supplier-based categorization
            if normalized_supplier in ["MedChemExpress", "ApexBio", "SelleckChem"]:
                return 'Drugs'
                # Check for supplier-specific categorization
            if normalized_supplier in biological_suppliers:
                return 'Biological'
            if normalized_supplier in mouse_suppliers:
                return 'Mouse Work'

            # Check name-based categorization
            for category, keywords in categories.items():
                if any(keyword.lower() in name_lower for keyword in keywords):
                    return category

            # Specific Category Matching
            specific_category_keywords = {
                'Antibodies': ['mouse antibody', 'rabbit antibody', 'goat antibody'],
                'Mouse Work': ['mouse cage', 'mouse bedding', 'animal cage']
            }
            for category, specific_keywords in specific_category_keywords.items():
                if any(keyword.lower() in name_lower for keyword in specific_keywords):
                    return category

            # General Category Matching
            for category, keywords in categories.items():
                if any(keyword.lower() in name_lower for keyword in keywords):
                    return category

            return 'Others'

        try:
            
            # Normalize column names to lowercase for consistency
            self.sheet_data.columns = [col.lower() for col in self.sheet_data.columns]

            # Step 4: Categorize each item based on name and supplier
            self.sheet_data['category'] = self.sheet_data.apply(
                lambda row: assign_category(row[name_column], row.get('supplier', '')), axis=1
            )

            # Ensure a 'cost' column exists (case-insensitive)
            if 'cost' not in self.sheet_data.columns:
                QMessageBox.warning(self.parent, "Missing Column", "The current sheet does not contain a 'Cost' column.")
                return

            # Step 5: Clean and convert the 'cost' column to numeric
            self.sheet_data['cost'] = self.sheet_data['cost'].replace(
                r'[^0-9.]', '', regex=True  # Remove all non-numeric characters except decimal points
            )
            
            # Replace empty strings or invalid values with 0 and convert to float
            self.sheet_data['cost'] = self.sheet_data['cost'].replace('', '0').astype(float)

            # Step 6: Group data by the 'category' column, calculate counts and total costs
            category_summary = self.sheet_data.groupby('category').agg(
                Count=('category', 'size'),
                Total_Cost=('cost', 'sum')
            ).reset_index()

            # Step 7: Add grouped data to a new sheet
            new_sheet_name = "Grouped_By_Category"
            counter = 1
            while new_sheet_name in self.sheet_dict:
                new_sheet_name = f"Grouped_By_Category_{counter}"
                counter += 1

            self.sheet_dict[new_sheet_name] = category_summary

            # Update UI to show the summary sheet
            table_widget = QTableWidget()
            table_widget.setRowCount(category_summary.shape[0])
            table_widget.setColumnCount(category_summary.shape[1])
            table_widget.setHorizontalHeaderLabels(category_summary.columns)

            for i, row in category_summary.iterrows():
                for j, value in enumerate(row):
                    table_widget.setItem(i, j, QTableWidgetItem(str(value)))

            table_widget.horizontalHeader().setSectionResizeMode(QHeaderView.Stretch)
            self.tab_widget.addTab(table_widget, new_sheet_name)
            self.tab_widget.setCurrentWidget(table_widget)

            QMessageBox.information(self.parent, "Success", "Category summary with total cost has been created in a new sheet.")

        except Exception as e:
            QMessageBox.critical(self.parent, "Error", f"An error occurred while categorizing items: {str(e)}")



    def allocate_costs_to_grant(self):
        """Allocate the selected costs to the selected grant."""
        selected_grant = self.grant_combo.currentText()
        selected_sum = float(re.sub(r'[^\d.]', '', self.selected_sum_label.text().split('$')[1]))

        grant_data = self.grant_management.get_grant_data(selected_grant)

        if grant_data is not None:
            total_grant_amount = grant_data['Total Balance'].iloc[0]

            # Safely access 'Allocated Costs' and handle cases where it may not exist
            if 'Allocated Costs' in grant_data.columns:
                allocated_cost = grant_data['Allocated Costs'].iloc[0]
            else:
                # Initialize 'Allocated Costs' to 0 if the column does not exist
                allocated_cost = 0
                self.grant_management.update_grant_data(selected_grant, 'Allocated Costs', allocated_cost)

            # Add the selected sum to the allocated costs
            updated_allocated_cost = allocated_cost + selected_sum
            net_amount = total_grant_amount - updated_allocated_cost

            # Update grant data with the new allocated costs and net amount
            self.grant_management.update_grant_data(selected_grant, 'Allocated Costs', updated_allocated_cost)
            self.grant_management.update_grant_data(selected_grant, 'Net Amount', net_amount)

            # Update the UI with the new net amount
            self.net_amount_label.setText(f"Net Amount in Grant: ${net_amount:.2f}")

            QMessageBox.information(self.parent, "Costs Allocated", f"Successfully allocated ${selected_sum:.2f} to the {selected_grant} grant.")
        else:
            QMessageBox.warning(self.parent, "Grant Not Found", f"The selected grant {selected_grant} could not be found.")

    def add_categorize_and_group_button(self):
        """
        Adds a button to categorize items in the current sheet and group them into a new sheet.
        """
        categorize_button = QPushButton("Categorize and Group Items")
        categorize_button.setStyleSheet("font-size: 16px; color: white; background-color: #4CAF50;")
        categorize_button.setFixedHeight(35)
        categorize_button.clicked.connect(self.categorize_and_group_items)
        self.right_button_layout.addWidget(categorize_button)

    def categorize_and_group_items(self):
        """
        Categorize items in the 'general + cost' sheet into predefined categories,
        ensuring no overlap between mouse antibodies and mouse work items.
        Supplier-specific categorization is also included.
        """
        if self.sheet_data is None:
            QMessageBox.warning(self.parent, "No Data", "No data available to categorize. Please load a sheet first.")
            return


        try:

            logging.basicConfig(level=logging.DEBUG)
            # Extended categories with keywords
            categories = {
                    'Media': [
                        # General media
                        'media', 'PBS', 'cell-culture', 'cell culture media', 'DMEM', 'RPMI', 'EMEM', 
                        'McCoy', 'IMDM', 'F-12', 'Ham\'s F-12', 'MEM', 'AMEM', 'α-MEM', 'Basal Medium Eagle',
                        'L-15', 'Leibovitz\'s L-15', 'Hank\'s Balanced Salt Solution', 'HBSS', 'Eagle\'s Medium', 
                        'Williams\' Medium E', 'Coon\'s Modified Ham\'s F-12', 'serum-free medium', 
                        'low-glucose medium', 'high-glucose medium', 'DMEM/F-12', 'RPMI-1640', 'keratinocyte medium',
                        
                        # Neutralizers and derivatives
                        'trypsin', 'trypsin-EDTA', 'trypsin neutralizer', 'TrypLE', 'trypsin substitute', 
                        'neutralizing solution', 'EDTA', 'collagenase', 'dispase', 'accutase', 
                        'cell dissociation solution', 'cell detachment solution', 'trypsin inhibitor',
                        
                        # Supplements and additives
                        'glutamine', 'L-glutamine', 'sodium pyruvate', 'non-essential amino acids', 
                        'NEAA', 'FBS', 'fetal bovine serum', 'bovine serum', 'horse serum', 'cell culture grade water',
                        'water for injection', 'sterile water', 'distilled water', 'di water', 'ultrapure water',
                        
                        # Specialized media
                        'neural stem cell medium', 'mesenchymal stem cell medium', 'embryonic stem cell medium',
                        'organoid culture media', 'hepatocyte media', 'airway epithelial cell media',
                        'fibroblast growth medium', 'skeletal muscle cell media', 'chondrocyte media', 
                        'endothelial growth medium', 'epithelial cell growth medium', 'keratinocyte serum-free medium',
                        
                        # Growth additives
                        'growth factor supplement', 'b27 supplement', 'N2 supplement', 'bfgf', 'EGF', 'insulin', 
                        'transferrin', 'selenium', 'hydrocortisone', 'dexamethasone', 'ascorbic acid', 'retinoic acid'
                    ],
                    'W/S/N Blots': [
                        'blot', 'western', 'southern', 'northern', 'gel', 'membrane', 'buffer', 'stain', 'substrate', 
                        'PAGE', 'SDS-PAGE', 'acrylamide', 'electrophoresis', 'ladder', 'marker', 'staining', 'mounting',
                        'HRP', 'chemiluminescent', 'chemiluminescence', 'fluorescent dye', 'immunoblot', 'immunoblotting', 
                        'transfer buffer', 'running buffer', 'blotting buffer', 'wash buffer', 'blocking buffer',
                        'PVDF', 'nitrocellulose', 'immobilon', 'BCA', 'Coomassie', 'silver stain', 'Ponceau', 
                        'NuPAGE', 'Bis-Tris', 'Tris-Glycine', 'MES buffer', 'MOPS buffer', 'transfer membrane',
                        'gel loading dye', 'protein ladder', 'DNA ladder', 'protein stain', 'anti-HRP', 'fluorescent marker',
                        'secondary detection', 'imaging substrate', 'ECL', 'enhanced chemiluminescence',
                        'polyacrylamide gel', 'Western substrate', 'Coomassie blue', 'chromogenic substrate',
                        'hybridization buffer', 'washing reagent', 'autoradiography', 'electroblotting', 'LDS Sample Buf', 'TBS w TWEEN TBST'
                    ],
                    'Antibodies': [
                        'antibody', 'antibodies', 'mAb', 'IgG', 'phospho', 'phospho-', 
                        'rabbit', 'mouse', 'goat', 'anti-', 'affinipure', 'monoclonal',
                        'secondary antibody', 'primary antibody', 'HRP-conjugated',
                        'Alexa Fluor', 'AF488', 'AF568', 'AF594', 'FITC', 'APC',
                        'Cy3', 'Cy5', 'Dylight', 'fluorescent antibody', 'polyclonal', 
                        'isotype control', 'conjugated antibody', 'biotinylated antibody', 
                        'peroxidase', 'HRP', 'AP (alkaline phosphatase)', 'ELISA antibody',
                        'immunoblot antibody', 'immunohistochemistry', 'IHC', 'ICC', 
                        'immunofluorescence', 'flow cytometry', 'western blot',

                        ## specific antibodies
                        'Hu Vimentin PE'
                    ],
                    'Flasks, Tips, etc.': [
                        'flask', 'flasks', 'Erlenmeyer flask', 'Erlenmeyer flasks', 'Conical flask', 'Conical flasks','well', 'wells',
                        'Cell culture flask', 'Cell culture flasks', 'Round-bottom flask', 'Round-bottom flasks',
                        'Volumetric flask', 'Volumetric flasks', 'Vacuum flask', 'Vacuum flasks', 'Filtering flask', 'Filtering flasks',
                        'tip', 'tips', 'pipet', 'pipets', 'pipette', 'pipettes', 'pipette tip', 'pipette tips',
                        'filter tip', 'filter tips', 'gel-loading tip', 'gel-loading tips', 'multi-channel tip', 'multi-channel tips',
                        'serological pipette', 'serological pipettes', 'manual pipette', 'manual pipettes', 'automatic pipette',
                        'automatic pipettes', 'multichannel pipette', 'multichannel pipettes', 'micropipette', 'micropipettes',
                        'repeater pipette', 'repeater pipettes', 'transfer pipette', 'transfer pipettes', 'glass pipette', 'glass pipettes',
                        'tube', 'tubes', 'centrifuge tube', 'centrifuge tubes', 'cryogenic tube', 'cryogenic tubes', 'chambers',
                        'microcentrifuge tube', 'microcentrifuge tubes', 'PCR tube', 'PCR tubes', 'glass tube', 'glass tubes',
                        'Falcon tube', 'Falcon tubes', 'Eppendorf tube', 'Eppendorf tubes', 'test tube', 'test tubes',
                        'storage tube', 'storage tubes', 'plts' ,'plate', 'plates', 'cell culture plate', 'cell culture plates',
                        'microplate', 'microplates', 'petri plate', 'petri plates', 'ELISA plate', 'ELISA plates',
                        'PCR plate', 'PCR plates', 'multi-well plate', 'multi-well plates', 'sealing plate', 'sealing plates',
                        'box', 'boxes', 'storage box', 'storage boxes', 'cryogenic box', 'cryogenic boxes', 'freezer box', 'freezer boxes',
                        'microtube box', 'microtube boxes', 'tip box', 'tip boxes', 'tube rack', 'tube racks',
                        'autoclave-safe box', 'autoclave-safe boxes', 'syringe', 'syringes', 'disposable syringe', 'disposable syringes',
                        'glass syringe', 'glass syringes', 'luer-lock syringe', 'luer-lock syringes', 'syringe filter', 'syringe filters',
                        'insulin syringe', 'insulin syringes', 'rack', 'racks', 'holder', 'holders', 'pipette rack', 'pipette racks',
                        'pipet rack', 'pipet racks', 'plate rack', 'plate racks', 'tube rack', 'tube racks', 'freezer rack', 'freezer racks',
                        'test tube rack', 'test tube racks', 'cryovial', 'cryovials', 'cryobox', 'cryoboxes', '384', 'allprotect tissue reagent',
                        'coutness', 'cryoelite', 'FBM', 'VWR BASIN'
                        'nitrogen storage rack', 'sterile container', 'sterile containers', 'sample vial', 'sample vials',
                        'funnel', 'funnels', 'glass slide', 'glass slides', 'coverslip', 'coverslips', 'weigh boat', 'weigh boats',
                        'measuring cylinder', 'measuring cylinders', 'spray bottle', 'spray bottles', 'lab tray', 'lab trays', 'T.I.P.S.',
                        'drip tray', 'drip trays', 'cell strainer', 'cell strainers', 'reservoir tray', 'reservoir trays', 'beaker', 'beakers', 'gloves', 'glove'
                    ],
                    'Assays': [
                        'assay', 'CyQUANT', 'DNeasy', 'Glo', 'immuno', 'ChIP', 'EdU', 'FITC', 'flow cytometry', 'mycoplasma', 'purelink hipure'
                    ],
                    'Mouse Work': [
                        'mouse', 'animal', 'rack', 'cage', 'rodent', 'bedding', 'scale', 'feeding', 'syringe for mouse', 
                        'mouse holder', 'animal cage'
                    ],
                    'Biological': [
                        # Existing items
                        'Lipofectamine', 'KAPA', 'concentrator', 'concentrators', 'goat serum', 'serum', 
                        'primers', 'primer', 'plasmid', 'glycerol stock', 'gBlock', 'lentivirus', 'Cas9', 'virus'
                        
                        # Enzymes and enzyme-related terms
                        'enzyme', 'restriction enzyme', 'ligase', 'polymerase', 'reverse transcriptase',
                        'DNA ligase', 'RNA polymerase', 'nuclease', 'endonuclease', 'exonuclease', 
                        'DNA polymerase', 'RNase', 'RNase inhibitor', 'phosphatase', 'kinase', 
                        'T4 ligase', 'Taq polymerase', 'Q5 polymerase', 'EcoRI', 'BamHI', 'NotI', 'HindIII',
                        'restriction digestion', 'digestion enzyme', 'proteinase K', 'Klenow fragment',
                        'DNase', 'DNAse I', 'methylase', 'NEBuilder', 'HiFi DNA Assembly', 'nickase',
                        
                        # NEB-specific items
                        'NEB', 'New England Biolabs', 'NEBuilder HiFi', 'Q5 Master Mix', 'Quick CIP',
                        'NEB ligase', 'NEB polymerase', 'NEB restriction enzyme', 'NEB buffer',
                        'NEBuffer', 'NEB T4 DNA Ligase', 'NEB Taq', 'NEB EcoRI', 'NEB digestion kit',
                        'NEB Phusion', 'NEB LunaScript', 'NEBNext', 'NEB methylase', 'NEB exonuclease',
                        
                        # Biological reagents and kits
                        'competent cells', 'cloning kit', 'transfection reagent', 'DNA assembly',
                        'electroporation reagent', 'viral vector', 'cDNA synthesis kit', 'PCR kit',
                        'RT-PCR kit', 'NGS prep kit', 'plasmid purification', 'protein ladder', 'SuperScript', 'RNeasy',
                        'marker', 'DNA ladder', 'RNA ladder', 'molecular weight marker', 'agarose', 'LB', 'agar',
                        
                        # Proteins and protein-related reagents
                        'protein expression', 'protein purification', 'proteinase', 'protease', 
                        'protein A', 'protein G', 'protein marker', 'protein standard', 'recombinant protein',
                        
                        # Cell culture additives and growth reagents
                        'cell culture reagent', 'cell growth reagent', 'supplement', 'cell recovery medium',
                        'freezing medium', 'cryopreservation', 'transfection reagent', 'nucleofection reagent',
                        
                        #Biologic Dyes
                        'Phalloidin', 'ANNEXIN V', 'ANNEXIN', 'HOECHST', 'vimentin live cell dye', 
                        'prolong diamond antifade mountant with dapi', 'dapi',

                        #Ladders
                        'GENERULER', 'master mix',

                        # Miscellaneous biological terms
                        'oligonucleotide', 'oligo', 'siRNA', 'shRNA', 'gRNA', 'sgRNA', 'RNAi', 'DNA template', 
                        'RNA template', 'expression plasmid', 'vector', 'CRISPR', 'CRISPR-Cas9', 'cloning vector',
                        'glycerol', 'competent cell', 'E.coli', 'BL21', 'DH5α', 'expression host', 'assembly mix', 'sequence',
                        'provirus', 'glucose', 'depc-treated'


                    ],
                    'Drugs': [
                        #Generic drugs
                        'drug', 'compound', 'chemical', 'inhibitor', 'small molecule', 'antibiotic', 'penicillin', 'amoxicillin', 'ciprofloxacin', 'azithromycin', 'cephalexin', 
                        'clindamycin', 'metronidazole', 'ampicillin', 'kanamycin', 'streptomycin', 'gentamicin', 'tetracycline', 'chloramphenicol', 
                        'penicillin', 'carbenicillin', 'antibiotic', 'small molecule', 'compound', 'chemical', 'inhibitor', 'aspirin', 'ibuprofen', 
                        'paracetamol', 'acetaminophen', 'statins', 'antiviral', 'aphidicolin', 'benzo(a)pyrene', 'doxycycline hyclate',
                        
                        #Antibiotics
                        'penicillin', 'streptomycin', 'ampicillin', 'kanamycin', 'tetracycline', 
                        'chloramphenicol', 'cephalosporin', 'erythromycin', 'rifampin', 'vancomycin', 
                        'gentamicin', 'ciprofloxacin', 'levofloxacin', 'azithromycin'

                        # Chemotherapy Drugs
                        'cisplatin', 'carboplatin', 'oxaliplatin', 'paclitaxel', 'docetaxel', 
                        'doxorubicin', 'epirubicin', 'cyclophosphamide', 'ifosfamide', 
                        'etoposide', 'irinotecan', 'topotecan', 'gemcitabine', 'vincristine', 
                        'vinblastine', 'vinorelbine', 'bleomycin', 'mitomycin', '5-fluorouracil', 
                        'capecitabine', 'methotrexate', 'pemetrexed', 'temozolomide', 'dacarbazine', 
                        'mechlorethamine', 'melphalan', 'busulfan', 'fludarabine', 'cladribine', 'OLAPARIB'],

                    'Chemical': [
                        'buffer', 'DMSO', 'ethanol', 'TCEP', 'methanol', 'glutaraldehyde', 
                        'SDS', 'Tris', 'HEPES', 'NaCl', 'TBE', 'formaldehyde', 'molecular biology', 'molecular'
                        'ammonia', 'glycine', 'crystal violet', 'ethyl cinnamate', 'iodonitrotetrazolium', 'Tetrakis(2-hydroxypropyl)',
                        'poly(ethylene glycol)', 'poly-l-lysine', 'protamine sulfate grade x', 'sulfo-smcc', 'tert-butanol'
                    ],
                    'Services (sequencing)': ['sequencing', 'service', 'genomics', 'WGS', 'long-read', 'sequencing service'],
                    'Services (one-time)': ['repair', 'installation', 'quote', 'service fee', 'one-time service', 'BSC'],
                    'Services (recurrent)': ['LN2', 'nitrogen', 'maintenance', 'subscription', 'recurring service', 'FY'],
                    'Office Supplies': ['ink cartridge', 'printer', 'stationery', 'WB Mason', 'pen', 'VWR Tape']
                
                    }

            # Original suppliers mapped to categories
            drug_suppliers = ['MedChemExpress', 'SelleckChem', 'ApexBio']
            enzyme_suppliers = ['New England Biolabs', 'NEB']
            plasmid_suppliers = ['Addgene', 'addgene']
            antibody_suppliers = ['Cell Signaling Technology']
            office_supplies_suppliers = ['WB Mason']
            mouse_suppliers = ['SoftMouse.NET','ISEEHEAR INC']
            biological_suppliers = ['Integrated DNA Technologies', 'VectorBuilder'] 


            # Normalization mapping for suppliers
            supplier_aliases = {
                'cell signaling': 'Cell Signaling Technology',
                'sigma aldrich': 'Millipore Sigma',
                'life tech': 'Life Technologies',
                'wb mason': 'WB Mason',
                'medchemexpress': 'MedChemExpress',
                'med chem express': 'MedChemExpress',
                'medchem express': 'MedChemExpress',
                'selleckchem': 'SelleckChem',
                'selleck chem': 'SelleckChem',
                'selleckchemicals': 'SelleckChem',
                'apexbio': 'ApexBio',
                'apex bio': 'ApexBio',
                'apexbiotechnology': 'ApexBio',
                'neb': 'New England Biolabs',
                'new england biolabs': 'New England Biolabs',
                'thermo fisher': 'Thermo Fisher',
                'invitrogen': 'Invitrogen',
                'promega': 'Promega',
                'bio-rad': 'Bio-Rad',
                'qiagen': 'Qiagen',
                'takara': 'Takara',
                'roche': 'Roche',
                'clontech': 'Clontech',
                'agilent': 'Agilent',
                'millipore': 'Millipore',
                'ge healthcare': 'GE Healthcare',
                'applied biosystems': 'Applied Biosystems',
                'epicentre': 'Epicentre',
                'softmouse.net': 'SoftMouse.NET',
                'ISEEHEAR INC': 'ISEEHEAR INC',
                'ISEEHEAR': 'ISEEHEAR INC',
                'idt': 'Integrated DNA Technologies', 
                'integrated dna technologies': 'Integrated DNA Technologies',
                'integrated dna tech': 'Integrated DNA Technologies',
                'vectorbuilder': 'VectorBuilder',  # Normalize VectorBuilder
                'vector builder': 'VectorBuilder'
            }


            # Step 1: Automatically find the 'name' column
            name_column = self.find_name_column()

            # Step 2: If not found, prompt the user to select the column
            if not name_column:
                column_names = self.sheet_data.columns.tolist()
                name_column, ok = QInputDialog.getItem(
                    self.parent, "Select Name Column",
                    "Choose the column containing item names:", column_names, 0, False
                )
                if not ok or not name_column:
                    QMessageBox.warning(self.parent, "Operation Cancelled", "No column was selected.")
                    return

            # Step 3: Ensure the name column exists
            if name_column not in self.sheet_data.columns:
                QMessageBox.warning(self.parent, "Missing Column", f"The current sheet does not contain a '{name_column}' column.")
                return

            # Function to normalize supplier names
            def normalize_supplier(supplier):
                if not supplier:
                    return None

                supplier_lower = supplier.lower().strip()

                # Check for known supplier patterns
                if "medchem" in supplier_lower:
                    return "MedChemExpress"
                if "apex" in supplier_lower:
                    return "ApexBio"
                if "selleck" in supplier_lower:
                    return "SelleckChem"
                if 'neb' in supplier_lower:
                    return 'New England Biolabs'
                if "idt" in supplier_lower or "integrated dna" in supplier_lower:
                    return "Integrated DNA Technologies"
                if "vectorbuilder" in supplier_lower or "vector builder" in supplier_lower:
                    return "VectorBuilder"

                # Return the original supplier if no match is found
                return supplier

            # Function to assign a category based on name and supplier
            def assign_category(name, supplier):
                name_lower = str(name).lower().strip()

                # Check if 'cisplatin' is in the name (prioritized)
                if "cisplatin" in name_lower:
                    return "Drugs"  # Prioritize 'cisplatin'


                # Normalize supplier
                normalized_supplier = normalize_supplier(supplier)

                    # Prioritized keyword matching
                if "gel ink pen" in name_lower or ("pen" in name_lower and "gel" in name_lower):
                    return "Office Supplies"  # Prioritize gel ink pens
                elif "western blot" in name_lower or ("gel" in name_lower and "blot" in name_lower):
                    return "W/S/N Blots"  # Secondary priority for western blot gels
                elif "gel" in name_lower:
                    # Additional logic for generic gels if needed
                    return "W/S/N Blots"  # Default to blot gels

                # Check supplier-based categorization
                if normalized_supplier in drug_suppliers:
                    return 'Drugs'
                if normalized_supplier in enzyme_suppliers:
                    return 'Biological'
                if normalized_supplier in plasmid_suppliers:
                    return 'Biological'
                if normalized_supplier in antibody_suppliers:
                    return 'Antibodies'
                if normalized_supplier in office_supplies_suppliers:
                    return 'Office Supplies'
                # Check supplier-based categorization
                if normalized_supplier in ["MedChemExpress", "ApexBio", "SelleckChem"]:
                    return 'Drugs'
                    # Check for supplier-specific categorization
                if normalized_supplier in biological_suppliers:
                    return 'Biological'
                if normalized_supplier in mouse_suppliers:
                    return 'Mouse Work'
    
                # Check name-based categorization
                for category, keywords in categories.items():
                    if any(keyword.lower() in name_lower for keyword in keywords):
                        return category

                # Specific Category Matching
                specific_category_keywords = {
                    'Antibodies': ['mouse antibody', 'rabbit antibody', 'goat antibody'],
                    'Mouse Work': ['mouse cage', 'mouse bedding', 'animal cage']
                }
                for category, specific_keywords in specific_category_keywords.items():
                    if any(keyword.lower() in name_lower for keyword in specific_keywords):
                        return category

                # General Category Matching
                for category, keywords in categories.items():
                    if any(keyword.lower() in name_lower for keyword in keywords):
                        return category

                return 'Others'

                # Step 4: Categorize each item based on name and supplier

            # Categorize items
            self.sheet_data['Category'] = self.sheet_data.apply(
                lambda row: assign_category(row[name_column], row.get('supplier', '')), axis=1
            )

            # Group data by the 'Category' column and count items
            grouped_data = self.sheet_data.groupby('Category').apply(
                lambda x: x.assign(Count=len(x))
            ).reset_index(drop=True)

            # Add grouped data as a new sheet
            new_sheet_name = "categorized_items"
            counter = 1
            while new_sheet_name in self.sheet_dict:
                new_sheet_name = f"categorized_items_{counter}"
                counter += 1

            self.sheet_dict[new_sheet_name] = grouped_data

            # Update UI to show the new categorized sheet
            table_widget = QTableWidget()
            table_widget.setRowCount(grouped_data.shape[0])
            table_widget.setColumnCount(grouped_data.shape[1])
            table_widget.setHorizontalHeaderLabels(grouped_data.columns)

            for i, row in grouped_data.iterrows():
                for j, value in enumerate(row):
                    table_widget.setItem(i, j, QTableWidgetItem(str(value)))

            table_widget.horizontalHeader().setSectionResizeMode(QHeaderView.Stretch)
            self.tab_widget.addTab(table_widget, new_sheet_name)
            self.tab_widget.setCurrentWidget(table_widget)

            QMessageBox.information(self.parent, "Success", "Items have been categorized and grouped into a new sheet.")

        except Exception as e:
            QMessageBox.critical(self.parent, "Error", f"An error occurred while categorizing items: {str(e)}")


    def display_saved_files(self):
        """Display a list of previously uploaded Excel files."""
        dialog = QDialog(self.parent)
        dialog.setWindowTitle("Previously Uploaded Excel Files")
        dialog.setStyleSheet("background-color: #cce7ff;")
        dialog.resize(400, 300)

        # Center the dialog on the screen
        screen = QApplication.primaryScreen().geometry()
        dialog_geometry = dialog.frameGeometry()
        dialog_geometry.moveCenter(screen.center())
        dialog.move(dialog_geometry.topLeft())

        layout = QVBoxLayout()

        file_list_widget = QListWidget()
        file_list_widget.setStyleSheet("font-size: 14px; color: #333; background-color: #f9f9f9;")

        # Populate the list widget with the saved file names
        for file_name in os.listdir(self.save_directory):
            if file_name.endswith(".xlsx"):
                file_list_widget.addItem(file_name)

        layout.addWidget(file_list_widget)

        open_button = QPushButton("Open Selected File")
        open_button.setStyleSheet("font-size: 16px; color: white; background-color: #4CAF50;")
        open_button.clicked.connect(lambda: self.open_selected_file(file_list_widget))
        layout.addWidget(open_button)

        dialog.setLayout(layout)
        dialog.exec_()

    def open_selected_file(self, file_list_widget):
        """Open and display the selected file from the list."""
        selected_file = file_list_widget.currentItem().text()

        if selected_file:
            file_path = os.path.join(self.save_directory, selected_file)
            try:
                # Load the selected file
                excel_data = pd.read_excel(file_path, sheet_name=None)
                self.display_excel_contents(excel_data)
            except Exception as e:
                QMessageBox.critical(self.parent, "Error", f"An error occurred while opening the Excel file: {str(e)}")
        else:
            QMessageBox.warning(self.parent, "No Selection", "Please select a file to open.")

